#!/usr/bin/env python3
"""
Security Report Generator for Wazuh Multi-Tenant Environment

Generates monthly security reports including:
1. Vulnerability Tracking (1st vs 30th comparison)
2. Security Incident Reporting (Level 12+ alerts)
3. Access Auditing (USA vs International, VPN activity)

USAGE:
  Run on 1st of month (or anytime) to capture baseline:
    python3 security_report.py --snapshot
  
  Run later in month to generate full report with comparison:
    python3 security_report.py
"""

import sys
import logging
import os
import json
import argparse
from datetime import datetime, timedelta
from typing import Dict, List, Any
from collections import defaultdict

import requests
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import Wazuh connector and Email reporter
from wazuh_connector import WazuhScanner
from email_reporter import EmailReporter

# Load environment variables from .env file
load_dotenv()

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SecurityReportGenerator:
    """Generate comprehensive security reports"""
    
    def __init__(self):
        self.config = self._load_config()
        self.wazuh = WazuhScanner(self.config['dashboard'])
        self.email_reporter = EmailReporter(
            smtp_server=self.config['email_server'],
            smtp_port=self.config['email_port'],
            sender=self.config['email_sender'],
            password=self.config['email_password']
        )
    
    def _load_config(self) -> Dict[str, Any]:
        """Load configuration from environment variables"""
        try:
            # Email configuration
            email_config = {
                'server': os.getenv('EMAIL_SERVER'),
                'port': int(os.getenv('EMAIL_PORT', 25)),
                'sender': os.getenv('EMAIL_SENDER'),
                'password': os.getenv('EMAIL_PASSWORD', ''),
            }
            
            # Wazuh dashboard configuration
            dashboard_config = {
                'host': os.getenv('WAZUH_HOST'),
                'port': int(os.getenv('WAZUH_PORT', 9200)),
                'username': os.getenv('WAZUH_USERNAME'),
                'password': os.getenv('WAZUH_PASSWORD'),
                'verify_ssl': os.getenv('WAZUH_VERIFY_SSL', 'false').lower() == 'true'
            }
            
            # Clients configuration
            clients_yaml = os.getenv('CLIENTS_CONFIG', 'clients: {}')
            clients_data = yaml.safe_load(clients_yaml) or {}
            
            return {
                'email_server': email_config['server'],
                'email_port': email_config['port'],
                'email_sender': email_config['sender'],
                'email_password': email_config['password'],
                'dashboard': dashboard_config,
                'clients': clients_data.get('clients', {})
            }
        except Exception as e:
            logger.error(f"Error loading configuration: {e}")
            sys.exit(1)
    
    def fetch_alerts_paginated(self, client_name: str, start_date: str, end_date: str, 
                                min_level: int = 12, batch_size: int = 10000) -> List[Dict]:
        """
        Fetch alerts with pagination (10k at a time)
        Excludes specific noise alerts
        """
        excluded_rules = [
            "Office 365: Phishing and malware events from Exchange Online Protection and Microsoft Defender for Office 365.",
            "Agent event queue is flooded. Check the agent configuration."
        ]
        
        all_alerts = []
        search_after = None
        
        must_not_clauses = [
            {"term": {"data.vulnerability.classification": "CVSS"}}
        ]
        for desc in excluded_rules:
            must_not_clauses.append({"match_phrase": {"rule.description": desc}})
        
        query = {
            "query": {
                "bool": {
                    "must": [
                        {"range": {"timestamp": {"gte": start_date, "lte": end_date}}},
                        {"range": {"rule.level": {"gte": min_level}}}
                    ],
                    "must_not": must_not_clauses
                }
            },
            "sort": [{"timestamp": "asc"}, {"_id": "asc"}],
            "size": batch_size
        }
        
        logger.info(f"Fetching alerts for {client_name} from {start_date} to {end_date}")
        
        while True:
            if search_after:
                query["search_after"] = search_after
            
            response = self.wazuh.search_client_alerts(client_name, query)
            
            if not response or 'hits' not in response or 'hits' not in response['hits']:
                break
            
            hits = response['hits']['hits']
            if not hits:
                break
            
            for hit in hits:
                alert = hit.get('_source', {})
                all_alerts.append(alert)
            
            # Get search_after value from last hit
            search_after = hits[-1].get('sort')
            
            logger.info(f"  Fetched {len(hits)} alerts (total: {len(all_alerts)})")
            
            # If we got less than batch_size, we're done
            if len(hits) < batch_size:
                break
        
        logger.info(f"Total alerts fetched: {len(all_alerts)}")
        return all_alerts
    
    def analyze_alerts(self, alerts: List[Dict]) -> Dict[str, Any]:
        """Analyze alerts for security incident reporting"""
        analysis = {
            'total_count': len(alerts),
            'by_level': defaultdict(int),
            'by_rule': defaultdict(int),
            'by_agent': defaultdict(int),
            'by_date': defaultdict(int),
            'top_rules': [],
            'unique_agents': set()
        }
        
        for alert in alerts:
            level = alert.get('rule', {}).get('level', 0)
            rule_id = alert.get('rule', {}).get('id', 'unknown')
            rule_desc = alert.get('rule', {}).get('description', 'Unknown')
            agent_name = alert.get('agent', {}).get('name', 'Unknown')
            timestamp = alert.get('timestamp', '')[:10]  # Get date only
            
            analysis['by_level'][f"Level {level}"] += 1
            analysis['by_rule'][f"{rule_id} - {rule_desc}"] += 1
            analysis['by_agent'][agent_name] += 1
            analysis['by_date'][timestamp] += 1
            analysis['unique_agents'].add(agent_name)
        
        # Get top 10 rules
        analysis['top_rules'] = sorted(
            analysis['by_rule'].items(), 
            key=lambda x: x[1], 
            reverse=True
        )[:10]
        
        analysis['unique_agents'] = list(analysis['unique_agents'])
        
        return analysis
    
    def analyze_access_auditing(self, alerts: List[Dict]) -> Dict[str, Any]:
        """Analyze login and VPN access patterns and extract detailed events"""
        access_analysis = {
            'usa_logins': 0,
            'international_logins': 0,
            'vpn_access': 0,
            'failed_logins': 0,
            'successful_logins': 0,
            'by_region': defaultdict(int),
            'by_user': defaultdict(int),
            'vpn_events': [],        # Store detailed VPN events
            'foreign_logins': [],    # Store detailed foreign login events
            'internal_logins': []    # Store detailed internal/local login events
        }
        
        # Keywords to identify login/VPN events
        login_keywords = ['login', 'logon', 'authentication', 'logged in', 'logon type']
        usa_indicators = ['United States', 'US', 'USA', 'America']
        
        for alert in alerts:
            rule_desc = alert.get('rule', {}).get('description', '').lower()
            data = alert.get('data', {})
            timestamp = alert.get('timestamp', '')
            
            # Check if it's a login event
            is_login = any(keyword in rule_desc for keyword in login_keywords)
            
            if is_login:
                # Check for internal Windows login (has win.eventdata fields)
                win_eventdata = data.get('win', {}).get('eventdata', {})
                target_user = win_eventdata.get('targetUserName', '')
                workstation = win_eventdata.get('workstationName', '')
                
                # If both internal fields exist, it's an internal login
                if target_user and workstation:
                    access_analysis['internal_logins'].append({
                        'timestamp': timestamp,
                        'user': target_user,
                        'workstation': workstation,
                        'source_ip': win_eventdata.get('ipAddress', ''),
                        'logon_type': win_eventdata.get('logonType', '')
                    })
                    continue  # Skip Office365 processing for internal logins
                
                # Get Office365 Scamalytics data
                scamalytics = data.get('Scamalytics', {})
                ipinfo = scamalytics.get('ipinfo', {})
                maxmind = scamalytics.get('maxmind', {})
                scamalytics_proxy = scamalytics.get('scamalytics_proxy', {})
                
                # Get location info
                country = ipinfo.get('country', '')
                city = maxmind.get('city', '')
                state = maxmind.get('state', '')
                
                # Get user from Office365 data
                office365 = data.get('office365', {})
                user = office365.get('UserId', 'Unknown')
                
                # Get source IP from Scamalytics
                source_ip = scamalytics.get('ip', '')
                
                # Get OS information
                os_info = data.get('device_properties', {}).get('OS', '')
                
                # Determine if VPN using Scamalytics proxy detection
                is_vpn = scamalytics_proxy.get('is_vpn', False)
                
                access_analysis['by_user'][user] += 1
                
                # VPN access event
                if is_vpn:
                    access_analysis['vpn_access'] += 1
                    
                    # Store detailed VPN event
                    access_analysis['vpn_events'].append({
                        'timestamp': timestamp,
                        'user': user,
                        'source_ip': source_ip,
                        'os': os_info
                    })
                
                # USA vs International (only for non-VPN logins)
                if country and not is_vpn:
                    is_usa = any(usa in country for usa in usa_indicators)
                    
                    if is_usa:
                        access_analysis['usa_logins'] += 1
                    else:
                        access_analysis['international_logins'] += 1
                        
                        # Store detailed foreign login event
                        access_analysis['foreign_logins'].append({
                            'timestamp': timestamp,
                            'user': user,
                            'source_ip': source_ip,
                            'country': country,
                            'city': city,
                            'state': state,
                            'os': os_info
                        })
                    
                    access_analysis['by_region'][country] += 1
                
                # Success vs Failure
                if 'fail' in rule_desc or 'denied' in rule_desc:
                    access_analysis['failed_logins'] += 1
                else:
                    access_analysis['successful_logins'] += 1
        
        # Sort events by timestamp (most recent first)
        access_analysis['vpn_events'] = sorted(
            access_analysis['vpn_events'], 
            key=lambda x: x['timestamp'], 
            reverse=True
        )
        
        access_analysis['foreign_logins'] = sorted(
            access_analysis['foreign_logins'], 
            key=lambda x: x['timestamp'], 
            reverse=True
        )
        
        access_analysis['internal_logins'] = sorted(
            access_analysis['internal_logins'], 
            key=lambda x: x['timestamp'], 
            reverse=True
        )
        
        return access_analysis
    
    def generate_excel_report(self, client_name: str, display_name: str,
                              alert_analysis: Dict, access_analysis: Dict,
                              vuln_start: Dict, vuln_current: Dict,
                              report_month: str, output_folder: str = ".") -> str:
        """Generate Excel report with vulnerability tracking"""
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{output_folder}/security_report_{client_name}_{report_month}_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # === SUMMARY SHEET ===
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="1F4E78")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, color="1F4E78", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws_summary['A1'] = "MONTHLY SECURITY REPORT"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        ws_summary.row_dimensions[1].height = 25
        
        # Report info
        row = 3
        ws_summary[f'A{row}'] = "Client:"
        ws_summary[f'B{row}'] = display_name.upper()
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Report Period:"
        ws_summary[f'B{row}'] = report_month
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Generated:"
        ws_summary[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        # Alert Summary
        row += 2
        ws_summary[f'A{row}'] = "SECURITY INCIDENT SUMMARY"
        ws_summary[f'A{row}'].font = section_font
        ws_summary[f'A{row}'].fill = section_fill
        ws_summary.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws_summary[f'A{row}'] = "Total Alerts (Level 12+):"
        ws_summary[f'B{row}'] = alert_analysis['total_count']
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Affected Systems:"
        ws_summary[f'B{row}'] = len(alert_analysis['unique_agents'])
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        # Vulnerability Summary
        row += 2
        ws_summary[f'A{row}'] = "VULNERABILITY TRACKING"
        ws_summary[f'A{row}'].font = section_font
        ws_summary[f'A{row}'].fill = section_fill
        ws_summary.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws_summary[f'A{row}'] = ""
        ws_summary[f'B{row}'] = "Month Start"
        ws_summary[f'C{row}'] = "Current"
        ws_summary[f'D{row}'] = "Change"
        for col in ['B', 'C', 'D']:
            ws_summary[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws_summary[f'{col}{row}'].fill = header_fill
        
        row += 1
        ws_summary[f'A{row}'] = "Total Vulnerabilities:"
        ws_summary[f'B{row}'] = vuln_start['total']
        ws_summary[f'C{row}'] = vuln_current['total']
        change = vuln_current['total'] - vuln_start['total']
        ws_summary[f'D{row}'] = f"{'+' if change > 0 else ''}{change}"
        ws_summary[f'D{row}'].font = Font(bold=True, color="FF0000" if change > 0 else "00B050")
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        # Add by severity
        for severity in ['Critical', 'High', 'Medium', 'Low']:
            row += 1
            start_count = vuln_start['by_severity'].get(severity, 0)
            current_count = vuln_current['by_severity'].get(severity, 0)
            change = current_count - start_count
            
            ws_summary[f'A{row}'] = f"  {severity}:"
            ws_summary[f'B{row}'] = start_count
            ws_summary[f'C{row}'] = current_count
            ws_summary[f'D{row}'] = f"{'+' if change > 0 else ''}{change}"
            ws_summary[f'D{row}'].font = Font(color="FF0000" if change > 0 else "00B050")
        
        # Access Auditing Summary
        row += 2
        ws_summary[f'A{row}'] = "ACCESS AUDITING SUMMARY"
        ws_summary[f'A{row}'].font = section_font
        ws_summary[f'A{row}'].fill = section_fill
        ws_summary.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws_summary[f'A{row}'] = "USA Logins:"
        ws_summary[f'B{row}'] = access_analysis['usa_logins']
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "International Logins:"
        ws_summary[f'B{row}'] = access_analysis['international_logins']
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "VPN Access Events:"
        ws_summary[f'B{row}'] = access_analysis['vpn_access']
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        # Set column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # === ALERTS BY LEVEL SHEET ===
        ws_levels = wb.create_sheet("Alerts by Level")
        ws_levels['A1'] = "Alert Level Distribution"
        ws_levels['A1'].font = title_font
        ws_levels.merge_cells('A1:B1')
        
        row = 3
        ws_levels['A3'] = "Level"
        ws_levels['B3'] = "Count"
        ws_levels['A3'].font = header_font
        ws_levels['A3'].fill = header_fill
        ws_levels['B3'].font = header_font
        ws_levels['B3'].fill = header_fill
        
        row = 4
        for level, count in sorted(alert_analysis['by_level'].items()):
            ws_levels[f'A{row}'] = level
            ws_levels[f'B{row}'] = count
            ws_levels[f'A{row}'].border = border
            ws_levels[f'B{row}'].border = border
            row += 1
        
        ws_levels.column_dimensions['A'].width = 20
        ws_levels.column_dimensions['B'].width = 15
        
        # === TOP RULES SHEET ===
        ws_rules = wb.create_sheet("Top Rules")
        ws_rules['A1'] = "Top 10 Triggered Rules"
        ws_rules['A1'].font = title_font
        ws_rules.merge_cells('A1:B1')
        
        ws_rules['A3'] = "Rule"
        ws_rules['B3'] = "Count"
        ws_rules['A3'].font = header_font
        ws_rules['A3'].fill = header_fill
        ws_rules['B3'].font = header_font
        ws_rules['B3'].fill = header_fill
        
        row = 4
        for rule, count in alert_analysis['top_rules']:
            ws_rules[f'A{row}'] = rule
            ws_rules[f'B{row}'] = count
            ws_rules[f'A{row}'].border = border
            ws_rules[f'B{row}'].border = border
            row += 1
        
        ws_rules.column_dimensions['A'].width = 80
        ws_rules.column_dimensions['B'].width = 15
        
        # === VULNERABILITY COMPARISON SHEET ===
        ws_vulns = wb.create_sheet("Vulnerability Comparison")
        
        # Title
        ws_vulns['A1'] = "Vulnerability Comparison (Month Start vs Current)"
        ws_vulns['A1'].font = title_font
        ws_vulns.merge_cells('A1:E1')
        
        # Headers
        ws_vulns['A3'] = "Agent"
        ws_vulns['B3'] = "Month Start"
        ws_vulns['C3'] = "Current"
        ws_vulns['D3'] = "Change"
        ws_vulns['E3'] = "Status"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_vulns[f'{col}3'].font = header_font
            ws_vulns[f'{col}3'].fill = header_fill
            ws_vulns[f'{col}3'].border = border
        
        # Get all unique agents
        all_agents = set(list(vuln_start['by_agent'].keys()) + list(vuln_current['by_agent'].keys()))
        
        row = 4
        for agent in sorted(all_agents):
            start_total = vuln_start['by_agent'].get(agent, {}).get('total', 0)
            current_total = vuln_current['by_agent'].get(agent, {}).get('total', 0)
            change = current_total - start_total
            
            ws_vulns[f'A{row}'] = agent
            ws_vulns[f'B{row}'] = start_total
            ws_vulns[f'C{row}'] = current_total
            ws_vulns[f'D{row}'] = f"{'+' if change > 0 else ''}{change}"
            
            if change > 0:
                ws_vulns[f'E{row}'] = "⬆ Increased"
                ws_vulns[f'E{row}'].font = Font(color="FF0000", bold=True)
            elif change < 0:
                ws_vulns[f'E{row}'] = "⬇ Decreased"
                ws_vulns[f'E{row}'].font = Font(color="00B050", bold=True)
            else:
                ws_vulns[f'E{row}'] = "→ No Change"
                ws_vulns[f'E{row}'].font = Font(color="808080")
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_vulns[f'{col}{row}'].border = border
            
            row += 1
        
        ws_vulns.column_dimensions['A'].width = 30
        ws_vulns.column_dimensions['B'].width = 15
        ws_vulns.column_dimensions['C'].width = 15
        ws_vulns.column_dimensions['D'].width = 15
        ws_vulns.column_dimensions['E'].width = 20
        
        # === VPN ACCESS SHEET ===
        ws_vpn = wb.create_sheet("VPN Access")
        
        # Title
        ws_vpn['A1'] = "VPN Access Events"
        ws_vpn['A1'].font = title_font
        ws_vpn.merge_cells('A1:D1')
        
        # Summary
        ws_vpn['A3'] = f"Total VPN Events: {access_analysis['vpn_access']}"
        ws_vpn['A3'].font = section_font
        
        # Headers
        ws_vpn['A5'] = "Timestamp"
        ws_vpn['B5'] = "User"
        ws_vpn['C5'] = "Source IP"
        ws_vpn['D5'] = "OS"
        
        for col in ['A', 'B', 'C', 'D']:
            ws_vpn[f'{col}5'].font = header_font
            ws_vpn[f'{col}5'].fill = header_fill
            ws_vpn[f'{col}5'].border = border
        
        # Add VPN events
        row = 6
        vpn_events = access_analysis.get('vpn_events', [])
        for event in vpn_events[:500]:  # Limit to 500 events
            ws_vpn[f'A{row}'] = event.get('timestamp', '')
            ws_vpn[f'B{row}'] = event.get('user', '')
            ws_vpn[f'C{row}'] = event.get('source_ip', '')
            ws_vpn[f'D{row}'] = event.get('os', '')
            
            for col in ['A', 'B', 'C', 'D']:
                ws_vpn[f'{col}{row}'].border = border
            
            row += 1
        
        ws_vpn.column_dimensions['A'].width = 20
        ws_vpn.column_dimensions['B'].width = 25
        ws_vpn.column_dimensions['C'].width = 18
        ws_vpn.column_dimensions['D'].width = 15
        
        # === FOREIGN LOGIN SHEET ===
        ws_foreign = wb.create_sheet("Foreign Logins")
        
        # Title
        ws_foreign['A1'] = "International Login Events"
        ws_foreign['A1'].font = title_font
        ws_foreign.merge_cells('A1:G1')
        
        # Summary
        ws_foreign['A3'] = f"Total International Logins: {access_analysis['international_logins']}"
        ws_foreign['A3'].font = section_font
        
        # Headers
        ws_foreign['A5'] = "Timestamp"
        ws_foreign['B5'] = "User"
        ws_foreign['C5'] = "Source IP"
        ws_foreign['D5'] = "Country"
        ws_foreign['E5'] = "City"
        ws_foreign['F5'] = "State"
        ws_foreign['G5'] = "OS"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_foreign[f'{col}5'].font = header_font
            ws_foreign[f'{col}5'].fill = header_fill
            ws_foreign[f'{col}5'].border = border
        
        # Add foreign login events
        row = 6
        foreign_events = access_analysis.get('foreign_logins', [])
        for event in foreign_events[:500]:  # Limit to 500 events
            ws_foreign[f'A{row}'] = event.get('timestamp', '')
            ws_foreign[f'B{row}'] = event.get('user', '')
            ws_foreign[f'C{row}'] = event.get('source_ip', '')
            ws_foreign[f'D{row}'] = event.get('country', '')
            ws_foreign[f'E{row}'] = event.get('city', '')
            ws_foreign[f'F{row}'] = event.get('state', '')
            ws_foreign[f'G{row}'] = event.get('os', '')
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_foreign[f'{col}{row}'].border = border
            
            row += 1
        
        ws_foreign.column_dimensions['A'].width = 20
        ws_foreign.column_dimensions['B'].width = 25
        ws_foreign.column_dimensions['C'].width = 18
        ws_foreign.column_dimensions['D'].width = 20
        ws_foreign.column_dimensions['E'].width = 18
        ws_foreign.column_dimensions['F'].width = 18
        ws_foreign.column_dimensions['G'].width = 15
        
        # === INTERNAL LOGIN SHEET ===
        ws_internal = wb.create_sheet("Internal Logins")
        
        # Title
        ws_internal['A1'] = "Internal/Local Login Events"
        ws_internal['A1'].font = title_font
        ws_internal.merge_cells('A1:E1')
        
        # Summary
        ws_internal['A3'] = f"Total Internal Logins: {len(access_analysis.get('internal_logins', []))}"
        ws_internal['A3'].font = section_font
        
        # Headers
        ws_internal['A5'] = "Timestamp"
        ws_internal['B5'] = "User"
        ws_internal['C5'] = "Workstation"
        ws_internal['D5'] = "Source IP"
        ws_internal['E5'] = "Logon Type"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_internal[f'{col}5'].font = header_font
            ws_internal[f'{col}5'].fill = header_fill
            ws_internal[f'{col}5'].border = border
        
        # Add internal login events
        row = 6
        internal_events = access_analysis.get('internal_logins', [])
        for event in internal_events[:500]:  # Limit to 500 events
            ws_internal[f'A{row}'] = event.get('timestamp', '')
            ws_internal[f'B{row}'] = event.get('user', '')
            ws_internal[f'C{row}'] = event.get('workstation', '')
            ws_internal[f'D{row}'] = event.get('source_ip', '')
            ws_internal[f'E{row}'] = event.get('logon_type', '')
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_internal[f'{col}{row}'].border = border
            
            row += 1
        
        ws_internal.column_dimensions['A'].width = 20
        ws_internal.column_dimensions['B'].width = 25
        ws_internal.column_dimensions['C'].width = 20
        ws_internal.column_dimensions['D'].width = 18
        ws_internal.column_dimensions['E'].width = 15
        
        # Save workbook
        wb.save(filename)
        logger.info(f"✅ Report saved: {filename}")
        
        return filename
    
    def fetch_vulnerabilities_snapshot(self, client_name: str, date: str) -> Dict[str, Any]:
        """Fetch vulnerability snapshot for a specific date
        
        Args:
            client_name: Client identifier
            date: Date in YYYY-MM-DD format
            
        Returns:
            Dictionary with vulnerability counts by severity and agent
        """
        severities = ["Critical", "High", "Medium", "Low"]
        snapshot = {
            'date': date,
            'total': 0,
            'by_severity': {},
            'by_agent': defaultdict(lambda: {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0, 'total': 0}),
            'vulnerabilities': []
        }
        
        for severity in severities:
            result = self.wazuh.search_vulnerabilities(client_name, severity, page_size=10000)
            
            if result.get('error'):
                logger.error(f"Error fetching {severity} vulnerabilities for {date}: {result['error']}")
                continue
            
            count = result.get('total', 0)
            snapshot['by_severity'][severity] = count
            snapshot['total'] += count
            
            # Track by agent
            for vuln in result.get('results', []):
                agent_name = vuln.get('agent', {}).get('name', 'Unknown')
                snapshot['by_agent'][agent_name][severity] += 1
                snapshot['by_agent'][agent_name]['total'] += 1
                
                # Store vulnerability details
                snapshot['vulnerabilities'].append({
                    'agent': agent_name,
                    'severity': severity,
                    'cve': vuln.get('vulnerability', {}).get('cve', 'N/A'),
                    'title': vuln.get('vulnerability', {}).get('title', 'Unknown'),
                    'package': vuln.get('package', {}).get('name', 'Unknown')
                })
        
        return snapshot
    
    def save_snapshot(self, client_name: str, snapshot: Dict[str, Any]):
        """Save vulnerability snapshot to JSON file for later comparison
        
        Only saves counts, not full vulnerability details, to keep files small.
        """
        snapshot_dir = f"reports/{client_name}/snapshots"
        os.makedirs(snapshot_dir, exist_ok=True)
        
        snapshot_file = f"{snapshot_dir}/snapshot_{snapshot['date']}.json"
        
        # Convert defaultdict to regular dict and ONLY save counts (no vulnerability details)
        snapshot_to_save = {
            'date': snapshot['date'],
            'total': snapshot['total'],
            'by_severity': snapshot['by_severity'],
            'by_agent': {agent: dict(counts) for agent, counts in snapshot['by_agent'].items()}
            # Note: 'vulnerabilities' list is NOT saved - only counts matter for comparison
        }
        
        with open(snapshot_file, 'w') as f:
            json.dump(snapshot_to_save, f, indent=2)
        
        logger.info(f"✅ Snapshot saved: {snapshot_file} ({snapshot['total']} vulnerabilities)")
        return snapshot_file
    
    def load_snapshot(self, client_name: str, date: str) -> Dict[str, Any]:
        """Load previously saved vulnerability snapshot"""
        snapshot_file = f"reports/{client_name}/snapshots/snapshot_{date}.json"
        
        if not os.path.exists(snapshot_file):
            logger.warning(f"No snapshot found for {date}: {snapshot_file}")
            return None
        
        with open(snapshot_file, 'r') as f:
            snapshot = json.load(f)
        
        logger.info(f"📂 Loaded snapshot from {date}: {snapshot['total']} vulnerabilities")
        return snapshot
    
    def capture_baseline_snapshot(self, client_name: str, client_config: Dict[str, Any]):
        """Capture and save baseline vulnerability snapshot (run on 1st of month)"""
        display_name = client_config.get('display_name', client_name.upper())
        now = datetime.now()
        today = now.strftime("%Y-%m-%d")
        
        logger.info("")
        logger.info("=" * 60)
        logger.info(f"Capturing Baseline Snapshot: {display_name}")
        logger.info(f"Date: {today}")
        logger.info("=" * 60)
        logger.info("")
        
        # Fetch current vulnerability state
        logger.info("Fetching vulnerability data...")
        snapshot = self.fetch_vulnerabilities_snapshot(client_name, today)
        
        # Save snapshot
        snapshot_file = self.save_snapshot(client_name, snapshot)
        
        logger.info(f"\n✅ Baseline captured: {snapshot['total']} vulnerabilities")
        logger.info(f"   Critical: {snapshot['by_severity'].get('Critical', 0)}")
        logger.info(f"   High: {snapshot['by_severity'].get('High', 0)}")
        logger.info(f"   Medium: {snapshot['by_severity'].get('Medium', 0)}")
        logger.info(f"   Low: {snapshot['by_severity'].get('Low', 0)}")
        logger.info("")

    def run_monthly_report(self, client_name: str, client_config: Dict[str, Any]):
        """Generate monthly security report for a specific client"""
        
        display_name = client_config.get('display_name', client_name.upper())
        
        # Get current month name
        now = datetime.now()
        report_month = now.strftime("%B %Y")
        year = now.year
        month = now.month
        
        logger.info("")
        logger.info("=" * 60)
        logger.info(f"Generating Security Report: {display_name}")
        logger.info(f"Period: {report_month}")
        logger.info("=" * 60)
        logger.info("")
        
        # Create client-specific output folder
        client_folder = f"reports/{client_name}"
        os.makedirs(client_folder, exist_ok=True)
        logger.info(f"Output folder: {client_folder}")
        
        # Try to load baseline snapshot from 1st of month
        baseline_date = f"{year}-{month:02d}-01"
        vuln_start = self.load_snapshot(client_name, baseline_date)
        
        if vuln_start is None:
            logger.warning(f"⚠️  No baseline snapshot found for {baseline_date}")
            logger.warning(f"   Run 'python3 security_report.py --snapshot' on the 1st to capture baseline")
            logger.info(f"   Using current data as baseline for this report...")
            vuln_start = self.fetch_vulnerabilities_snapshot(client_name, baseline_date)
            self.save_snapshot(client_name, vuln_start)
        
        # Fetch current vulnerability state
        logger.info("Fetching current vulnerability data...")
        vuln_current = self.fetch_vulnerabilities_snapshot(client_name, now.strftime("%Y-%m-%d"))
        
        # Fetch alerts from start of month to now
        start_date = f"{year}-{month:02d}-01T00:00:00"
        end_date = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        
        # Fetch high-level security alerts (Level 12+) for incident reporting
        logger.info(f"Fetching high-level alerts (Level 12+) for {client_name} from {start_date} to {end_date}")
        alerts = self.fetch_alerts_paginated(client_name, start_date, end_date, min_level=12)
        
        # Fetch ALL login/authentication events (Level 3+) for access auditing
        logger.info(f"Fetching login events (Level 3+) for {client_name}...")
        login_alerts = self.fetch_alerts_paginated(client_name, start_date, end_date, min_level=3)
        
        if not alerts and not login_alerts and vuln_start['total'] == 0 and vuln_current['total'] == 0:
            logger.warning(f"No data found for {client_name}")
            return
        
        # Analyze alerts
        logger.info("Analyzing security incidents...")
        alert_analysis = self.analyze_alerts(alerts) if alerts else {'total_count': 0, 'by_level': {}, 'top_rules': [], 'unique_agents': []}
        
        logger.info("Analyzing access patterns...")
        access_analysis = self.analyze_access_auditing(login_alerts) if login_alerts else {
            'usa_logins': 0, 'international_logins': 0, 'vpn_access': 0,
            'failed_logins': 0, 'successful_logins': 0, 'by_region': {}, 'by_user': {},
            'vpn_events': [], 'foreign_logins': [], 'internal_logins': []
        }
        
        # Generate Excel report with vulnerability comparison
        logger.info("Generating Excel report...")
        report_file = self.generate_excel_report(
            client_name=client_name,
            display_name=display_name,
            alert_analysis=alert_analysis,
            access_analysis=access_analysis,
            vuln_start=vuln_start,
            vuln_current=vuln_current,
            report_month=report_month,
            output_folder=client_folder
        )
        
        # Send email
        email_recipients = client_config.get('email_recipients', [])
        if email_recipients:
            logger.info("Sending report via email...")
            for recipient in email_recipients:
                try:
                    self.email_reporter.send_report(
                        recipient=recipient,
                        client_name=display_name,
                        excel_file=report_file
                    )
                    logger.info(f"✅ Report sent to {recipient}")
                except Exception as e:
                    logger.error(f"Failed to send to {recipient}: {e}")
        
        logger.info(f"\n✅ Security report generation complete for {client_name}\n")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Security Report Generator for Wazuh Multi-Tenant',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Usage Examples:
  
  1. Capture baseline snapshot (run on 1st of month):
     python3 security_report.py --snapshot
  
  2. Generate full report with comparison (run anytime after baseline):
     python3 security_report.py
  
  3. Generate report for specific client only:
     python3 security_report.py --client homelab
  
  4. Capture baseline for specific client:
     python3 security_report.py --snapshot --client homelab

Workflow:
  Step 1: On the 1st (or beginning) of month, run with --snapshot to save baseline
  Step 2: Later in month (or end of month), run without --snapshot to generate report
          The report will compare baseline vs current vulnerabilities
        """
    )
    
    parser.add_argument(
        '--snapshot',
        action='store_true',
        help='Capture baseline vulnerability snapshot only (no report generation)'
    )
    
    parser.add_argument(
        '--client',
        type=str,
        help='Process specific client only (default: all enabled clients)'
    )
    
    args = parser.parse_args()
    
    generator = SecurityReportGenerator()
    
    # Determine which clients to process
    if args.client:
        if args.client not in generator.config['clients']:
            logger.error(f"Client '{args.client}' not found in configuration")
            sys.exit(1)
        clients_to_process = {args.client: generator.config['clients'][args.client]}
    else:
        clients_to_process = generator.config['clients']
    
    # Process clients
    for client_name, client_config in clients_to_process.items():
        if not client_config.get('enabled', False):
            logger.info(f"Client {client_name} is disabled, skipping")
            continue
        
        try:
            if args.snapshot:
                # Capture baseline snapshot only
                generator.capture_baseline_snapshot(client_name, client_config)
            else:
                # Generate full report
                generator.run_monthly_report(client_name, client_config)
        except Exception as e:
            logger.error(f"Error processing {client_name}: {e}")
            import traceback
            traceback.print_exc()
            continue


if __name__ == '__main__':
    main()
