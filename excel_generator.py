#!/usr/bin/env python3
"""
Excel Report Generator
Generates formatted Excel reports with multiple sheets
"""

import logging
from datetime import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class SecurityExcelGenerator:
    """Generate formatted Excel security reports"""
    
    # Predefined styles
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
    SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SECTION_FONT = Font(bold=True, color="1F4E78", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def generate_report(self, client_name: str, display_name: str,
                       alert_analysis: Dict, access_analysis: Dict,
                       vuln_start: Dict, vuln_current: Dict,
                       report_month: str, output_folder: str = ".") -> str:
        """Generate comprehensive Excel report
        
        Args:
            client_name: Client identifier (for filename)
            display_name: Client display name (for report title)
            alert_analysis: Alert analysis results
            access_analysis: Access pattern analysis results
            vuln_start: Baseline vulnerability snapshot
            vuln_current: Current vulnerability snapshot
            report_month: Report period (e.g., "November 2025")
            output_folder: Output directory path
            
        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{output_folder}/security_report_{client_name}_{report_month}_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Create all sheets
        self._create_summary_sheet(wb.active, display_name, report_month, 
                                   alert_analysis, access_analysis, vuln_start, vuln_current)
        self._create_alerts_by_level_sheet(wb, alert_analysis)
        self._create_top_rules_sheet(wb, alert_analysis)
        self._create_vulnerability_comparison_sheet(wb, vuln_start, vuln_current)
        self._create_access_events_sheet(wb, "VPN Access", access_analysis['vpn_events'], 
                                        ['Timestamp', 'User', 'Source IP', 'OS'],
                                        ['timestamp', 'user', 'source_ip', 'os'],
                                        access_analysis['vpn_access'])
        self._create_access_events_sheet(wb, "Foreign Logins", access_analysis['foreign_logins'],
                                        ['Timestamp', 'User', 'Source IP', 'Country', 'City', 'State', 'OS'],
                                        ['timestamp', 'user', 'source_ip', 'country', 'city', 'state', 'os'],
                                        access_analysis['international_logins'])
        self._create_access_events_sheet(wb, "Internal Logins", access_analysis['internal_logins'],
                                        ['Timestamp', 'User', 'Workstation', 'Source IP', 'Logon Type'],
                                        ['timestamp', 'user', 'workstation', 'source_ip', 'logon_type'],
                                        len(access_analysis['internal_logins']))
        
        wb.save(filename)
        logger.info(f" Report saved: {filename}")
        return filename
    
    def _create_summary_sheet(self, ws, display_name, report_month, alert_analysis, 
                             access_analysis, vuln_start, vuln_current):
        """Create summary sheet with key metrics"""
        ws.title = "Summary"
        
        # Title
        ws['A1'] = "MONTHLY SECURITY REPORT"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        # Report info
        row = 3
        info = [
            ("Client:", display_name.upper()),
            ("Report Period:", report_month),
            ("Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Alert Summary
        row += 1
        ws[f'A{row}'] = "SECURITY INCIDENT SUMMARY"
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        summary_items = [
            ("Total Alerts (Level 12+):", alert_analysis['total_count']),
            ("Affected Systems:", len(alert_analysis['unique_agents']))
        ]
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Vulnerability Summary
        row += 1
        ws[f'A{row}'] = "VULNERABILITY TRACKING"
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for col, header in [('B', 'Month Start'), ('C', 'Current'), ('D', 'Change')]:
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = self.HEADER_FONT
            ws[f'{col}{row}'].fill = self.HEADER_FILL
        
        row += 1
        ws[f'A{row}'] = "Total Vulnerabilities:"
        ws[f'B{row}'] = vuln_start['total']
        ws[f'C{row}'] = vuln_current['total']
        change = vuln_current['total'] - vuln_start['total']
        ws[f'D{row}'] = f"{'+' if change > 0 else ''}{change}"
        ws[f'D{row}'].font = Font(bold=True, color="FF0000" if change > 0 else "00B050")
        ws[f'A{row}'].font = Font(bold=True)
        
        # By severity
        for severity in ['Critical', 'High', 'Medium', 'Low']:
            row += 1
            start_count = vuln_start['by_severity'].get(severity, 0)
            current_count = vuln_current['by_severity'].get(severity, 0)
            change = current_count - start_count
            
            ws[f'A{row}'] = f"  {severity}:"
            ws[f'B{row}'] = start_count
            ws[f'C{row}'] = current_count
            ws[f'D{row}'] = f"{'+' if change > 0 else ''}{change}"
            ws[f'D{row}'].font = Font(color="FF0000" if change > 0 else "00B050")
        
        # Access Auditing Summary
        row += 2
        ws[f'A{row}'] = "ACCESS AUDITING SUMMARY"
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        access_items = [
            ("International Logins:", access_analysis['international_logins']),
            ("VPN Access Events:", access_analysis['vpn_access'])
        ]
        for label, value in access_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def _create_alerts_by_level_sheet(self, wb, alert_analysis):
        """Create alerts by level distribution sheet"""
        ws = wb.create_sheet("Alerts by Level")
        ws['A1'] = "Alert Level Distribution"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "Level"
        ws['B3'] = "Count"
        for col in ['A3', 'B3']:
            ws[col].font = self.HEADER_FONT
            ws[col].fill = self.HEADER_FILL
        
        row = 4
        for level, count in sorted(alert_analysis['by_level'].items()):
            ws[f'A{row}'] = level
            ws[f'B{row}'] = count
            ws[f'A{row}'].border = self.BORDER
            ws[f'B{row}'].border = self.BORDER
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_top_rules_sheet(self, wb, alert_analysis):
        """Create top triggered rules sheet"""
        ws = wb.create_sheet("Top Rules")
        ws['A1'] = "Top 10 Triggered Rules"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "Rule"
        ws['B3'] = "Count"
        for col in ['A3', 'B3']:
            ws[col].font = self.HEADER_FONT
            ws[col].fill = self.HEADER_FILL
        
        row = 4
        for rule, count in alert_analysis['top_rules']:
            ws[f'A{row}'] = rule
            ws[f'B{row}'] = count
            ws[f'A{row}'].border = self.BORDER
            ws[f'B{row}'].border = self.BORDER
            row += 1
        
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 15
    
    def _create_vulnerability_comparison_sheet(self, wb, vuln_start, vuln_current):
        """Create vulnerability comparison sheet"""
        ws = wb.create_sheet("Vulnerability Comparison")
        
        ws['A1'] = "Vulnerability Comparison (Month Start vs Current)"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:E1')
        
        headers = ["Agent", "Month Start", "Current", "Change", "Status"]
        for col_idx, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_idx)
            ws[f'{col_letter}3'] = header
            ws[f'{col_letter}3'].font = self.HEADER_FONT
            ws[f'{col_letter}3'].fill = self.HEADER_FILL
            ws[f'{col_letter}3'].border = self.BORDER
        
        # Get all unique agents
        all_agents = set(list(vuln_start.get('by_agent', {}).keys()) + 
                        list(vuln_current.get('by_agent', {}).keys()))
        
        row = 4
        for agent in sorted(all_agents):
            start_total = vuln_start.get('by_agent', {}).get(agent, {}).get('total', 0)
            current_total = vuln_current.get('by_agent', {}).get(agent, {}).get('total', 0)
            change = current_total - start_total
            
            ws[f'A{row}'] = agent
            ws[f'B{row}'] = start_total
            ws[f'C{row}'] = current_total
            ws[f'D{row}'] = f"{'+' if change > 0 else ''}{change}"
            
            if change > 0:
                ws[f'E{row}'] = "⬆ Increased"
                ws[f'E{row}'].font = Font(color="FF0000", bold=True)
            elif change < 0:
                ws[f'E{row}'] = "⬇ Decreased"
                ws[f'E{row}'].font = Font(color="00B050", bold=True)
            else:
                ws[f'E{row}'] = "→ No Change"
                ws[f'E{row}'].font = Font(color="808080")
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = self.BORDER
            
            row += 1
        
        widths = [30, 15, 15, 15, 20]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width
    
    def _create_access_events_sheet(self, wb, sheet_name, events, headers, fields, total_count):
        """Create generic access events sheet (VPN, Foreign, Internal logins)
        
        Args:
            wb: Workbook object
            sheet_name: Name of the sheet
            events: List of event dictionaries
            headers: List of column headers
            fields: List of field names to extract from events
            total_count: Total count for summary
        """
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = f"{sheet_name} Events"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells(f'A1:{chr(64 + len(headers))}1')
        
        # Summary
        ws['A3'] = f"Total {sheet_name}: {total_count}"
        ws['A3'].font = self.SECTION_FONT
        
        # Headers
        for col_idx, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_idx)
            ws[f'{col_letter}5'] = header
            ws[f'{col_letter}5'].font = self.HEADER_FONT
            ws[f'{col_letter}5'].fill = self.HEADER_FILL
            ws[f'{col_letter}5'].border = self.BORDER
        
        # Add events (limit to 500)
        row = 6
        for event in events[:500]:
            for col_idx, field in enumerate(fields, start=1):
                col_letter = chr(64 + col_idx)
                ws[f'{col_letter}{row}'] = event.get(field, '')
                ws[f'{col_letter}{row}'].border = self.BORDER
            row += 1
        
        # Set column widths
        default_widths = {
            'Timestamp': 20, 'User': 25, 'Source IP': 18, 'OS': 15,
            'Country': 20, 'City': 18, 'State': 18, 'Workstation': 20,
            'Logon Type': 15
        }
        for col_idx, header in enumerate(headers, start=1):
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = default_widths.get(header, 20)